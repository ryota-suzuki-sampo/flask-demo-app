from flask import Flask, render_template, request, redirect, url_for, send_file, flash, jsonify
from flask_login import LoginManager, login_user, logout_user, login_required, UserMixin, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import os
import psycopg2
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl import load_workbook
import logging

app = Flask(__name__, template_folder="templates")
app.secret_key = os.environ.get("SECRET_KEY", "secret-key")

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

# --- DB接続 ---
def get_conn():
    return psycopg2.connect(os.environ["DATABASE_URL"])

# --- User クラス ---
class User(UserMixin):
    def __init__(self, id, username, password_hash):
        self.id = id
        self.username = username
        self.password_hash = password_hash

    @staticmethod
    def get_by_username(username):
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id, username, password_hash FROM users WHERE username = %s", (username,))
                row = cur.fetchone()
                if row:
                    return User(*row)
        return None

    @staticmethod
    def get(user_id):
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id, username, password_hash FROM users WHERE id = %s", (user_id,))
                row = cur.fetchone()
                if row:
                    return User(*row)
        return None

@login_manager.user_loader
def load_user(user_id):
    return User.get(user_id)

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        user = User.get_by_username(username)
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            return redirect(url_for("list_ships"))
        flash("ユーザー名またはパスワードが正しくありません")
    return render_template("login.html")

@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("ログアウトしました")
    return redirect(url_for("login"))

@app.route("/")
def home_redirect():
    return redirect("/ships")

@app.route("/register", methods=["GET", "POST"])
@login_required
def register():
    if request.method == "POST":
        ship_name = request.form["ship_name"]
        company_name = request.form["company_name"]
        charter_type = request.form["charter_type"]
        flag = request.form["flag"]
        ship_type = request.form["ship_type"]
        completion_date = request.form["completion_date"]

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO ships (ship_name, company_name, charter_type, flag, ship_type, completion_date)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (ship_name, company_name, charter_type, flag, ship_type, completion_date))
        return redirect("/ships")

    return render_template("register.html")

@app.route("/ships")
@login_required
def list_ships():
    search = request.args.get("search", "")
    sort = request.args.get("sort", "id")
    order = request.args.get("order", "desc")
    page = int(request.args.get("page", 1))
    per_page = 10
    offset = (page - 1) * per_page

    allowed_sorts = ["id", "ship_name", "company_name", "completion_date"]
    allowed_orders = ["asc", "desc"]
    sort = sort if sort in allowed_sorts else "id"
    order = order if order in allowed_orders else "desc"

    with get_conn() as conn:
        with conn.cursor() as cur:
            if search:
                cur.execute("""
                    SELECT COUNT(*) FROM ships
                    WHERE ship_name ILIKE %s OR company_name ILIKE %s
                """, (f"%{search}%", f"%{search}%"))
            else:
                cur.execute("SELECT COUNT(*) FROM ships")
            total = cur.fetchone()[0]
            total_pages = (total + per_page - 1) // per_page

            if search:
                cur.execute(f"""
                    SELECT id, ship_name, company_name, charter_type, completion_date, flag, ship_type
                    FROM ships
                    WHERE ship_name ILIKE %s OR company_name ILIKE %s
                    ORDER BY {sort} {order}
                    LIMIT {per_page} OFFSET {offset}
                """, (f"%{search}%", f"%{search}%"))
            else:
                cur.execute(f"""
                    SELECT id, ship_name, company_name, charter_type, completion_date, flag, ship_type
                    FROM ships
                    ORDER BY {sort} {order}
                    LIMIT {per_page} OFFSET {offset}
                """)
            ships = cur.fetchall()

    return render_template("ships.html", ships=ships, search=search, sort=sort, order=order, page=page, total_pages=total_pages)

@app.route("/ships/<int:ship_id>")
def ship_detail(ship_id):
    edit_mode = request.args.get("edit") == "1"

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT ship_name FROM ships WHERE id = %s", (ship_id,))
            ship = cur.fetchone()
            if not ship:
                return "Not Found", 404

            cur.execute("SELECT id, name FROM currencies")
            currencies = cur.fetchall()

            cur.execute("""
                SELECT charter_currency_id, charter_fee,
                    ship_currency_id, ship_cost,
                    repayment_currency_id, repayment,
                    interest_currency_id, interest,
                    loan_balance_currency_id, loan_balance,
                    fx_reserve_currency_id, fx_reserve_amount, fx_reserve_rate
                FROM ship_details
                WHERE ship_id = %s
            """, (ship_id,))
            detail = cur.fetchone()

            # 🚨 安全に interest を %表示変換（=100倍）
            if detail:
                detail = list(detail)
                if len(detail) >= 8 and detail[7] is not None:
                    detail[7] = round(detail[7] * 100, 2)

    return render_template("ship_detail.html",
                           ship_id=ship_id,
                           ship_name=ship[0],
                           currencies=currencies,
                           detail=detail,
                           edit=edit_mode)


@app.route("/ships/<int:ship_id>/update", methods=["POST"])
def update_ship_detail(ship_id):
    interest_input = request.form.get("interest")
    interest = float(interest_input) / 100 if interest_input else None

    data = {
        "charter_currency_id": request.form.get("charter_currency_id"),
        "charter_fee": request.form.get("charter_fee"),
        "ship_currency_id": request.form.get("ship_currency_id"),
        "ship_cost": request.form.get("ship_cost"),
        "repayment_currency_id": request.form.get("repayment_currency_id"),
        "repayment": request.form.get("repayment"),
        "interest_currency_id": request.form.get("interest_currency_id"),
        "interest": interest,
        "loan_balance_currency_id": request.form.get("loan_balance_currency_id"),
        "loan_balance": request.form.get("loan_balance"),
        "fx_reserve_currency_id": request.form.get("fx_reserve_currency_id"),
        "fx_reserve_amount": request.form.get("fx_reserve_amount"),
        "fx_reserve_rate": request.form.get("fx_reserve_rate")
    }

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM ship_details WHERE ship_id = %s", (ship_id,))
            exists = cur.fetchone()
            if exists:
                cur.execute("""
                    UPDATE ship_details
                    SET charter_currency_id = %s, charter_fee = %s,
                        ship_currency_id = %s, ship_cost = %s,
                        repayment_currency_id = %s, repayment = %s,
                        interest_currency_id = %s, interest = %s,
                        loan_balance_currency_id = %s, loan_balance = %s,
                        fx_reserve_currency_id = %s, fx_reserve_amount = %s, fx_reserve_rate = %s
                    WHERE ship_id = %s
                """, (*data.values(), ship_id))
            else:
                cur.execute("""
                    INSERT INTO ship_details (
                        ship_id,
                        charter_currency_id, charter_fee,
                        ship_currency_id, ship_cost,
                        repayment_currency_id, repayment,
                        interest_currency_id, interest, loan_balance_currency_id, loan_balance ,
                        fx_reserve_currency_id, fx_reserve_amount, fx_reserve_rate
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (ship_id, *data.values()))

    return redirect(url_for("ship_detail", ship_id=ship_id))

@app.route("/export_excel", methods=["POST"])
def export_excel():
    ship_ids = request.form.getlist("ship_ids")
    template_file = request.files.get("template_file")

    if not ship_ids or not template_file:
        return "船舶選択とテンプレートファイルの両方が必要です", 400

    # テンプレート読み込み
    wb = openpyxl.load_workbook(template_file)
    if "format" not in wb.sheetnames:
        return "テンプレートに 'format' シートが存在しません", 400

    # 新シート作成
    now_str = datetime.now().strftime("%Y%m%d%H%M")
    ws_template = wb["format"]
    ws_output = wb.copy_worksheet(ws_template)
    ws_output.title = f"Output_{now_str}"
    ws_template.sheet_view.tabSelected = False
    wb.active = wb.index(ws_output)

    # DBから対象データ取得
    with get_conn() as conn:
        with conn.cursor() as cur:
            format_ids = tuple(map(int, ship_ids))
            cur.execute(f"""
                SELECT s.id, s.ship_name,
                       cd1.name AS charter_currency, sd.charter_fee,
                       cd2.name AS ship_currency, sd.ship_cost,
                       cd3.name AS repayment_currency, sd.repayment,
                       cd4.name AS interest_currency, sd.interest,
                       cd5.name AS loan_currency, sd.loan_balance
                FROM ships s
                LEFT JOIN ship_details sd ON s.id = sd.ship_id
                LEFT JOIN currencies cd1 ON sd.charter_currency_id = cd1.id
                LEFT JOIN currencies cd2 ON sd.ship_currency_id = cd2.id
                LEFT JOIN currencies cd3 ON sd.repayment_currency_id = cd3.id
                LEFT JOIN currencies cd4 ON sd.interest_currency_id = cd4.id
                LEFT JOIN currencies cd5 ON sd.loan_balance_currency_id = cd5.id
                WHERE s.id IN %s
                ORDER BY s.id
            """, (format_ids,))
            records = cur.fetchall()

    # 4行目から書き込み
    start_row = 4
    for idx, row in enumerate(records, start=1):
        ws_output.cell(row=start_row, column=2, value=idx)                    # B列: 連番
        ws_output.cell(row=start_row, column=3, value=row[1])                 # C列: 船名
        ws_output.cell(row=start_row, column=4, value=row[2])                 # D列: 傭船料通貨
        ws_output.cell(row=start_row, column=5, value=row[3])                 # E列: 傭船料金額
        ws_output.cell(row=start_row, column=6, value=row[4])                 # F列: 船舶費通貨
        ws_output.cell(row=start_row, column=7, value=row[5])                 # G列: 船舶費金額
        ws_output.cell(row=start_row, column=8, value=row[6])                 # H列: 元利金通貨
        ws_output.cell(row=start_row, column=9, value=row[7])                 # I列: 元利金金額

        # J列: 利息（%表示）
        if row[9] is not None:
            cell = ws_output.cell(row=start_row, column=10, value=row[9])

        start_row += 1

    # 出力ファイルをバッファに保存
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"ShipExport_{now_str}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/change_password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "POST":
        current_pw = request.form["current_password"]
        new_pw = request.form["new_password"]
        confirm_pw = request.form["confirm_password"]

        user = current_user

        if not check_password_hash(user.password_hash, current_pw):
            flash("現在のパスワードが正しくありません")
        elif new_pw != confirm_pw:
            flash("新しいパスワードが一致しません")
        else:
            new_hash = generate_password_hash(new_pw)
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("UPDATE users SET password_hash = %s WHERE id = %s", (new_hash, user.id))
            flash("パスワードを変更しました")
            return redirect(url_for("list_ships"))

    return render_template("change_password.html")

@app.route('/aggregate_start', methods=['GET'])
@login_required
def aggregate_start():
    now = datetime.now()
    return render_template('aggregate_start.html', now=now)

@app.route('/api/ship_names', methods=['POST'])
@login_required
def api_ship_names():
    ship_ids = request.json.get('ship_ids', [])
    if not ship_ids:
        return jsonify([])
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT ship_name FROM ships WHERE id = ANY(%s) ORDER BY id",
                (list(map(int, ship_ids)),)
            )
            rows = cur.fetchall()
    return jsonify([r[0] for r in rows])
# --- 追加：Excel出力設定 ---
EXPORT_CONFIG = {
    'start_month': 'E4',
    'usd_range_cols': list(range(5, 17)),  # E〜P列

    'charter_usd_row': 6,
    'cost_usd_row': 10,
    'cost_spec_row': 24,
    'repay_usd_row': 13,
    'repay_spec_row': 32,
    'interest_usd_row': 16,
    'interest_spec_row': 35,
    'loan_usd_cell': (17, 4),    # D17
    'loan_spec_cell': (36, 4),   # D36
    'shipname_start_cell': (5, 19),  # S5〜
    'fx_reserve_row': 12,       # 為替予約金額
    'fx_reserve_yen_row': 47,   # 金額換算（円）
}

def write_values(ws, row, cols, value):
    """同一行の複数列に同じ値を代入"""
    for col in cols:
        ws.cell(row=row, column=col).value = value

@app.route('/export_aggregated_excel', methods=['POST'])
@login_required
def export_aggregated_excel():
    # フォームデータ取得
    start_month   = request.form['start_month']
    template_file = request.files['template_file']
    ship_ids      = request.form.getlist('ship_ids')

    if not ship_ids:
        return redirect(url_for('aggregate_start'))

    ids = list(map(int, ship_ids))

    # 1) 傭船料合計
    sql_charter = """
        SELECT cd.name AS currency, COALESCE(SUM(sd.charter_fee), 0) AS total
          FROM ship_details sd
          JOIN currencies cd ON sd.charter_currency_id = cd.id
         WHERE sd.ship_id = ANY(%s)
         GROUP BY cd.name
    """
    # 2) 船舶費合計
    sql_cost = """
        SELECT cd.name AS currency, COALESCE(SUM(sd.ship_cost), 0) AS total
          FROM ship_details sd
          JOIN currencies cd ON sd.ship_currency_id = cd.id
         WHERE sd.ship_id = ANY(%s)
         GROUP BY cd.name
    """
    # 3) 返済額合計
    sql_repay = """
        SELECT cd.name AS currency, COALESCE(SUM(sd.repayment), 0) AS total
          FROM ship_details sd
          JOIN currencies cd ON sd.repayment_currency_id = cd.id
         WHERE sd.ship_id = ANY(%s)
         GROUP BY cd.name
    """
    # 4) 支払利息平均（小数→パーセントに変換）
    sql_interest = """
        SELECT cd.name AS currency, AVG(sd.interest) AS avg_val
          FROM ship_details sd
          JOIN currencies cd ON sd.interest_currency_id = cd.id
         WHERE sd.ship_id = ANY(%s)
         GROUP BY cd.name
    """
    # 5) 融資残高合計
    sql_loan = """
        SELECT cd.name AS currency, COALESCE(SUM(sd.loan_balance), 0) AS total
          FROM ship_details sd
          JOIN currencies cd ON sd.loan_balance_currency_id = cd.id
         WHERE sd.ship_id = ANY(%s)
         GROUP BY cd.name
    """
    # 6) 船舶名一覧取得
    sql_names = """
        SELECT ship_name
          FROM ships
         WHERE id = ANY(%s)
         ORDER BY id
    """
    # 7) 為替予約情報
    sql_fx_reserve = """
        SELECT cd.name AS currency,
               COALESCE(SUM(sd.fx_reserve_amount), 0) AS total_amount,
               COALESCE(AVG(sd.fx_reserve_rate), 0) AS avg_rate
          FROM ship_details sd
          JOIN currencies cd ON sd.fx_reserve_currency_id = cd.id
         WHERE sd.ship_id = ANY(%s)
         GROUP BY cd.name
    """

    # データ取得
    charter_totals = {}
    cost_totals    = {}
    repay_totals   = {}
    interest_avgs  = {}
    loan_totals    = {}
    ship_names     = []
    fx_reserve_data = {}

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql_charter,  (ids,))
            charter_totals = dict(cur.fetchall())
            print("CHARTER:", charter_totals)
            app.logger.debug(f"CHARTER: {charter_totals}")

            cur.execute(sql_cost,     (ids,))
            cost_totals    = dict(cur.fetchall())
            print("COST:", cost_totals)

            cur.execute(sql_repay,    (ids,))
            repay_totals   = dict(cur.fetchall())
            print("REPAY:", repay_totals)

            cur.execute(sql_interest, (ids,))
            interest_avgs  = dict(cur.fetchall())
            print("INTEREST AVG:", interest_avgs)

            cur.execute(sql_loan,     (ids,))
            loan_totals    = dict(cur.fetchall())
            print("LOAN:", loan_totals)

            cur.execute(sql_names,    (ids,))
            ship_names = [r[0] for r in cur.fetchall()]
            print("SHIP NAMES:", ship_names)

            cur.execute(sql_fx_reserve, (ids,))
            fx_reserve_data = {
                row[0]: {
                    'amount': row[1],
                    'rate': row[2]
                }
                for row in cur.fetchall()
            }
            print("FX RESERVE:", fx_reserve_data)
            app.logger.info("FX RESERVE: %s", fx_reserve_data)

    # Excelテンプレート読み込み
    wb = load_workbook(template_file.stream)
    buf = BytesIO()

    # 「収支合計_預金管理_XXX」シートの候補
    valid_codes = ['JPY', 'CHF', 'XEU']

    # 返済通貨ごとにシートを選択し書き込み
    for code, repay_val in repay_totals.items():
        sheet_name = f"収支合計_{code}"
        if code not in valid_codes or sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        config = EXPORT_CONFIG

        # 開始年月
        ws[config['start_month']] = start_month

        # 傭船料（USD）
        write_values(ws, config['charter_usd_row'], config['usd_range_cols'], charter_totals.get('USD', 0))

        # 船舶費（USD / 指定通貨）
        write_values(ws, config['cost_usd_row'], config['usd_range_cols'], cost_totals.get('USD', 0))
        write_values(ws, config['cost_spec_row'], config['usd_range_cols'], cost_totals.get(code, 0))

        # 返済額（USD / 指定通貨）
        # write_values(ws, config['repay_usd_row'], config['usd_range_cols'], repay_totals.get('USD', 0))
        write_values(ws, config['repay_spec_row'], config['usd_range_cols'], repay_val)

        # 支払利息（USD / 指定通貨）
        # write_values(ws, config['interest_usd_row'], config['usd_range_cols'], interest_avgs.get('USD', 0))
        write_values(ws, config['interest_spec_row'], config['usd_range_cols'], interest_avgs.get(code, 0))

        # 融資残高（USD / 指定通貨）
        # ws.cell(*config['loan_usd_cell'], value=loan_totals.get('USD', 0))
        ws.cell(*config['loan_spec_cell'], value=loan_totals.get(code, 0))

        # 為替予約情報
        fx_data = fx_reserve_data.get(code, {'amount': 0, 'rate': 0})
        fx_amount = fx_data['amount']
        fx_rate   = fx_data['rate']
        fx_yen    = fx_amount * fx_rate

        # 12ヶ月分展開
        write_values(ws, config['fx_reserve_row'], config['usd_range_cols'], fx_amount)
        write_values(ws, config['fx_reserve_yen_row'], config['usd_range_cols'], fx_yen)

        # 船舶名リスト出力（S列40行目から）
        r, col = config['shipname_start_cell']
        for name in ship_names:
            ws.cell(row=r, column=col, value=name)
            r += 1

    # 保存して返却
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=template_file.filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def write_usd_detail_sheet(ws, ship_list, charter_by_ship, cost_by_ship, loan_by_ship, repay_by_ship, interest_by_ship, ship_name_by_id):
    row = 6
    for ship_id in ship_list:
        ship_name = ship_name_by_id.get(ship_id)
        if not ship_name:
            continue

        charter = charter_by_ship.get(ship_id, 0)
        cost = cost_by_ship.get(ship_id, 0)
        loan = loan_by_ship.get(ship_id, 0)
        repay = repay_by_ship.get(ship_id, 0)
        interest = interest_by_ship.get(ship_id, 0)

        if loan == 0 or repay == 0:
            continue

        repayment_monthly = repay / 12
        monthly_days = 30  # 1か月の暦日（必要に応じて調整）
        balance = loan
        interest_total = 0
        balance_sum = 0

        for m in range(12):
            monthly_interest = balance * interest * monthly_days / 365
            interest_total += monthly_interest
            balance_sum += balance
            balance -= repayment_monthly

        average_balance = balance_sum / 12

        # 書き込み
        ws[f'B{row}'] = ship_name
        ws[f'C{row}'] = charter * 365
        ws[f'E{row}'] = cost * 12
        ws[f'F{row}'] = loan
        ws[f'G{row}'] = average_balance
        ws[f'H{row}'] = repay
        ws[f'J{row}'] = interest_total

        row += 1

@app.route('/export_2currency_aggregated_excel', methods=['POST'])
@login_required
def export_2currency_aggregated_excel():
    # フォームデータ取得
    start_month   = request.form['start_month']
    template_file = request.files['template_file']
    ship_ids      = request.form.getlist('ship_ids')

    print("export_2currency_aggregated_excel Start")

    if not ship_ids:
        print("export_2currency_aggregated_excel not ship end")
        return redirect(url_for('aggregate_start'))

    ids = list(map(int, ship_ids))
    # --- 1) 返済通貨を船単位で取得 ---
    sql_ship_repay_currency = """
        SELECT sci.ship_id, c.name AS repayment_currency
        FROM ship_cost_items sci
        JOIN cost_item_type_table cit ON sci.item_type_id = cit.id
        JOIN currencies c ON sci.currency_id = c.id
        WHERE cit.item_code = 'repayment'
        AND sci.ship_id = ANY(%s)
        GROUP BY sci.ship_id, c.name
    """
    # 1) 傭船料合計
    sql_ship_charter = """
        SELECT sci.ship_id, COALESCE(SUM(sci.amount), 0) AS total
        FROM ship_cost_items sci
        JOIN cost_item_type_table cit ON sci.item_type_id = cit.id
        JOIN currencies c ON sci.currency_id = c.id
        WHERE cit.item_code = 'charter'
          AND c.name = 'USD'
          AND sci.ship_id = ANY(%s)
        GROUP BY sci.ship_id
    """
    # 2) 船舶費合計（通貨単位）
    sql_cost = """
        SELECT c.name AS currency, COALESCE(SUM(sci.amount), 0) AS total
        FROM ship_cost_items sci
        JOIN cost_item_type_table cit ON sci.item_type_id = cit.id
        JOIN currencies c ON sci.currency_id = c.id
        WHERE cit.item_code = 'ship'
          AND sci.ship_id = ANY(%s)
        GROUP BY c.name
    """
    # 2-2) 船舶費合計（船単位）
    sql_ship_cost = """
        SELECT sci.ship_id, COALESCE(SUM(sci.amount), 0) AS total
        FROM ship_cost_items sci
        JOIN cost_item_type_table cit ON sci.item_type_id = cit.id
        JOIN currencies c ON sci.currency_id = c.id
        WHERE cit.item_code = 'ship'
          AND sci.ship_id = ANY(%s)
        GROUP BY sci.ship_id
    """
    # 3) 返済額合計
    sql_repay = """
        SELECT c.name AS currency, COALESCE(SUM(sci.amount), 0) AS total
        FROM ship_cost_items sci
        JOIN cost_item_type_table cit ON sci.item_type_id = cit.id
        JOIN currencies c ON sci.currency_id = c.id
        WHERE cit.item_code = 'repayment'
          AND sci.ship_id = ANY(%s)
        GROUP BY c.name
    """
    # 3-2) 返済額合計（船単位）
    sql_ship_repay = """
        SELECT sci.ship_id, COALESCE(SUM(sci.amount), 0) AS total
        FROM ship_cost_items sci
        JOIN cost_item_type_table cit ON sci.item_type_id = cit.id
        JOIN currencies c ON sci.currency_id = c.id
        WHERE cit.item_code = 'repayment'
          AND sci.ship_id = ANY(%s)
        GROUP BY sci.ship_id
    """
    # 4) 支払利息平均（通貨単位）
    sql_interest = """
        SELECT c.name AS currency, AVG(sci.amount) AS avg_val
        FROM ship_cost_items sci
        JOIN cost_item_type_table cit ON sci.item_type_id = cit.id
        JOIN currencies c ON sci.currency_id = c.id
        WHERE cit.item_code = 'interest'
          AND sci.ship_id = ANY(%s)
        GROUP BY c.name
    """
    # 4-2) 支払利息平均（船単位）
    sql_ship_interest = """
        SELECT sci.ship_id, AVG(sci.amount) AS avg_val
        FROM ship_cost_items sci
        JOIN cost_item_type_table cit ON sci.item_type_id = cit.id
        JOIN currencies c ON sci.currency_id = c.id
        WHERE cit.item_code = 'interest'
          AND sci.ship_id = ANY(%s)
        GROUP BY sci.ship_id
    """
    # 5) 融資残高合計（通貨単位）
    sql_loan = """
        SELECT c.name AS currency, COALESCE(SUM(sci.amount), 0) AS total
        FROM ship_cost_items sci
        JOIN cost_item_type_table cit ON sci.item_type_id = cit.id
        JOIN currencies c ON sci.currency_id = c.id
        WHERE cit.item_code = 'loan'
          AND sci.ship_id = ANY(%s)
        GROUP BY c.name
    """
    # 5-2) 融資残高合計（船単位)
    sql_ship_loan = """
        SELECT sci.ship_id, COALESCE(SUM(sci.amount), 0) AS total
        FROM ship_cost_items sci
        JOIN cost_item_type_table cit ON sci.item_type_id = cit.id
        JOIN currencies c ON sci.currency_id = c.id
        WHERE cit.item_code = 'loan'
          AND sci.ship_id = ANY(%s)
        GROUP BY sci.ship_id
    """
    # 6) 船舶名一覧取得
    sql_names = """
        SELECT id, ship_name
        FROM ships
        WHERE id = ANY(%s)
        ORDER BY id
    """
    # 7) 為替予約情報
    #sql_fx_reserve = """
    #    SELECT cd.name AS currency,
    #           COALESCE(SUM(sd.fx_reserve_amount), 0) AS total_amount,
    #           COALESCE(AVG(sd.fx_reserve_rate), 0) AS avg_rate
    ##      FROM ship_details sd
    #      JOIN currencies cd ON sd.fx_reserve_currency_id = cd.id
    #     WHERE sd.ship_id = ANY(%s)
    #     GROUP BY cd.name
    #"""

    # データ取得
    charter_totals = {}
    cost_totals    = {}
    repay_totals   = {}
    interest_avgs  = {}
    loan_totals    = {}
    ship_names     = []
    #fx_reserve_data = {}

    print("export_2currency_aggregated_excel SQL")

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql_ship_repay_currency, (ids,))
            repay_currency_by_ship = dict(cur.fetchall())
            cur.execute(sql_ship_charter, (ids,))
            charter_by_ship = dict(cur.fetchall())
            # --- repayment 通貨単位で合計 ---
            charter_totals = {}
            for ship_id, amount in charter_by_ship.items():
                repay_currency = repay_currency_by_ship.get(ship_id)
                if not repay_currency:
                    continue
                charter_totals[repay_currency] = charter_totals.get(repay_currency, 0) + amount
            print("CHARTER:", charter_totals)

            cur.execute(sql_cost,     (ids,))
            cost_totals = dict(cur.fetchall())
            print("COST:", cost_totals)

            cur.execute(sql_repay,    (ids,))
            repay_totals = dict(cur.fetchall())
            print("REPAY:", repay_totals)

            cur.execute(sql_interest, (ids,))
            interest_avgs = dict(cur.fetchall())
            print("INTEREST AVG:", interest_avgs)

            cur.execute(sql_loan,     (ids,))
            loan_totals = dict(cur.fetchall())
            print("LOAN:", loan_totals)

            cur.execute(sql_names,    (ids,))
            ship_name_dict = dict(cur.fetchall())
            print("SHIP NAMES:", ship_names)

            cur.execute(sql_ship_cost, (ids,))
            cost_by_ship = dict(cur.fetchall())
            cur.execute(sql_ship_loan, (ids,))
            loan_by_ship = dict(cur.fetchall())
            cur.execute(sql_ship_repay, (ids,))
            repay_by_ship = dict(cur.fetchall())
            cur.execute(sql_ship_interest, (ids,))
            interest_by_ship = dict(cur.fetchall())
#            cur.execute(sql_fx_reserve, (ids,))
#            fx_reserve_data = {
#                row[0]: {
#                    'amount': row[1],
#                    'rate': row[2]
#                }
#                for row in cur.fetchall()
#            }
#            print("FX RESERVE:", fx_reserve_data)
#            app.logger.info("FX RESERVE: %s", fx_reserve_data)

    # Excelテンプレート読み込み
    print("export_2currency_aggregated_excel load Excel File")
    wb = load_workbook(template_file.stream)
    buf = BytesIO()

    # 「収支合計_預金管理_XXX」シートの候補
    valid_codes = ['JPY', 'CHF', 'XEU', 'USD']

    # 返済通貨ごとにシートを選択し書き込み
    for code, repay_val in repay_totals.items():
        detail_sheet = None
        if code == 'USD':
            sheet_name = "収支合計_金利_USD"
            detail_sheet = "金利_USD"
        else:
            sheet_name = f"収支合計_為替_{code}"

        if code not in valid_codes or sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        config = EXPORT_CONFIG

        # 開始年月
        ws[config['start_month']] = start_month

        # 傭船料（USD）
        write_values(ws, config['charter_usd_row'], config['usd_range_cols'], charter_totals.get(code, 0))

        # 船舶費（USD / 指定通貨）
        write_values(ws, config['cost_usd_row'], config['usd_range_cols'], cost_totals.get('USD', 0))
        if sheet_name == f"収支合計_為替_{code}":
            write_values(ws, config['cost_spec_row'], config['usd_range_cols'], cost_totals.get(code, 0))

        # 返済額（USD / 指定通貨）
        write_values(ws, config['repay_usd_row'], config['usd_range_cols'], repay_totals.get('USD', 0))
        write_values(ws, config['repay_spec_row'], config['usd_range_cols'], repay_val)

        # 支払利息（USD / 指定通貨）
        write_values(ws, config['interest_usd_row'], config['usd_range_cols'], interest_avgs.get('USD', 0))
        write_values(ws, config['interest_spec_row'], config['usd_range_cols'], interest_avgs.get(code, 0))

        # 融資残高（USD / 指定通貨）
        ws.cell(*config['loan_usd_cell'], value=loan_totals.get('USD', 0))
        ws.cell(*config['loan_spec_cell'], value=loan_totals.get(code, 0))

        # 為替予約情報
        #fx_data = fx_reserve_data.get(code, {'amount': 0, 'rate': 0})
        #fx_amount = fx_data['amount']
        #fx_rate   = fx_data['rate']
        #fx_yen    = fx_amount * fx_rate

        # 12ヶ月分展開
        #write_values(ws, config['fx_reserve_row'], config['usd_range_cols'], fx_amount)
        #write_values(ws, config['fx_reserve_yen_row'], config['usd_range_cols'], fx_yen)

        # 船舶名リスト出力（S列40行目から）
        r, col = config['shipname_start_cell']
        for name in ship_name_dict.values():
            ws.cell(row=r, column=col, value=name)
            r += 1

        if detail_sheet and detail_sheet in wb.sheetnames:
            ws_detail = wb[detail_sheet]
            # 呼び出し関数に必要情報を渡す
            usd_ship_ids = [sid for sid, curr in repay_currency_by_ship.items() if curr == 'USD']
            write_usd_detail_sheet(ws_detail, usd_ship_ids, charter_by_ship, cost_by_ship, loan_by_ship, repay_by_ship, interest_by_ship, ship_name_dict)

    # 保存して返却
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=template_file.filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route("/ships/<int:ship_id>/cost_items", methods=["GET", "POST"])
@login_required
def manage_cost_items(ship_id):
    with get_conn() as conn:
        with conn.cursor() as cur:
            # 船名取得
            cur.execute("SELECT ship_name FROM ships WHERE id = %s", (ship_id,))
            row = cur.fetchone()
            if not row:
                return "Not Found", 404
            ship_name = row[0]

            # マスタ取得
            cur.execute("SELECT id, name FROM currencies ORDER BY id")
            currencies = cur.fetchall()
            cur.execute("SELECT id, item_name, unit FROM cost_item_type_table ORDER BY id")
            item_types = cur.fetchall()

            if request.method == "POST":
                # 一度削除してから再INSERT（簡易処理）
                cur.execute("DELETE FROM ship_cost_items WHERE ship_id = %s", (ship_id,))
                loan_balance_currency_map = {}
                for item_id in [i[0] for i in item_types]:
                    for gno in [1, 2]:
                        currency = request.form.get(f"currency_{item_id}_{gno}")
                        amount = request.form.get(f"amount_{item_id}_{gno}")

                        if item_id == 5 and currency:
                            loan_balance_currency_map[gno] = currency
                        if item_id == 6:
                            ratio = request.form.get(f"ratio_{item_id}_{gno}")
                            currency = loan_balance_currency_map.get(gno)
                            amount = ratio

                        if currency and amount:
                            try:
                                amount_val = float(amount.replace(',', '')) if isinstance(amount, str) else float(amount)
                                print("item_id : ", item_id)
                                if item_id == 4:  # 支払利息（％入力 → 実数保存）
                                    print("amount_val1 : ", amount_val)
                                    amount_val /= 100
                                    print("amount_val2 : ", amount_val)
                                cur.execute("""
                                    INSERT INTO ship_cost_items
                                    (ship_id, item_type_id, group_no, currency_id, amount)
                                    VALUES (%s, %s, %s, %s, %s)
                                """, (ship_id, item_id, gno, currency, amount_val))
                            except ValueError:
                                pass
                return redirect(url_for("manage_cost_items", ship_id=ship_id))

            # GET: 既存データ読み出し
            cur.execute("""
                SELECT item_type_id, group_no, currency_id, amount
                  FROM ship_cost_items
                 WHERE ship_id = %s
            """, (ship_id,))
            rows = cur.fetchall()

            # 辞書化（item_type_id → {group_no → {currency_id, amount}})
            cost_data = {}
            for item_id, gno, curid, amt in rows:
                cost_data.setdefault(item_id, {})[gno] = {
                    "currency_id": curid,
                    "ratio" if item_id == 6 else "amount": float(amt)
                }

    return render_template("ship_cost_items.html",
                           ship_id=ship_id,
                           ship_name=ship_name,
                           item_types=item_types,
                           currencies=currencies,
                           cost_data=cost_data)

if __name__ == "__main__":
    print("Starting app on port 5000...")
    app.logger.setLevel(logging.DEBUG)
    app.run(host="0.0.0.0", port=5000)
